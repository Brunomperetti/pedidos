import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
import unicodedata
import requests
import tempfile
import math
from io import BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from io import BytesIO

# --- Funciones base de tu c√≥digo ---

def quitar_acentos(texto: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFKD", str(texto))
        if not unicodedata.combining(c)
    ).lower()

@st.cache_data(show_spinner=False)
def fetch_excel_from_drive(file_id: str) -> Path:
    url = f"https://drive.google.com/uc?export=download&id={file_id}"
    r = requests.get(url, timeout=30)
    r.raise_for_status()
    tmp = Path(tempfile.gettempdir()) / f"{file_id}.xlsx"
    tmp.write_bytes(r.content)
    return tmp

@st.cache_data(show_spinner=False)
def load_products_from_sheet(xls_path: str, sheet_name: str) -> pd.DataFrame:
    wb = load_workbook(xls_path, data_only=True)
    ws = wb[sheet_name]  # Cargar la hoja espec√≠fica
    
    # Leer las filas a partir de la fila 3 (suponiendo que las primeras filas son encabezados)
    rows = []
    images = {}  # Diccionario para almacenar las im√°genes

    for idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if not row[0]:  # Si no hay SKU, se detiene
            break
        # Extraer solo las columnas necesarias (SKU, Imagen, Descripci√≥n, etc.)
        sku, imagen, descripcion, tama√±o, precio_usd, unidades_por_caja, *resto = row[:7]

        # Buscar im√°genes embebidas en la hoja
        for img in ws._images:
            if img.anchor._from.col == 1 and img.anchor._from.row == idx - 1:  # Relacionar con la fila
                img_bytes = img._data()
                images[sku] = img_bytes  # Almacenar la imagen en un diccionario con SKU como clave
        
        # Filtramos solo las columnas necesarias
        rows.append([sku, imagen, descripcion, tama√±o, precio_usd, unidades_por_caja])

    # Crear DataFrame con las columnas especificadas
    df = pd.DataFrame(rows, columns=[
        "SKU", "Imagen", "Descripcion", "Tama√±o del producto", "Precio USD", "Unidades por caja"
    ])
    
    # Limpiar las filas con datos incompletos si es necesario
    df = df.dropna(subset=["SKU", "Descripcion", "Precio USD", "Unidades por caja"])  # Eliminar filas con datos faltantes en columnas cr√≠ticas
    
    # Normalizar los datos
    df["descripcion_norm"] = df["Descripcion"].apply(quitar_acentos)
    
    # Limpiar precios (eliminando signos de d√≥lar y comas)
    df["Precio USD"] = df["Precio USD"].apply(lambda x: float(str(x).replace("$", "").replace(",", "")) if isinstance(x, str) else x)
    
    return df, images

# --- Variables principales ---

st.set_page_config(page_title="Cat√°logo Millex", layout="wide")

# ID del archivo de Google Sheets
FILE_ID = "1JG-_vjmFXnWM13Xp6PCOzjgJkxks8BEF"  # Reemplazar con el ID de tu archivo

# Cargar el archivo Excel desde Google Drive
xls_path = fetch_excel_from_drive(FILE_ID)

# Cargar las hojas disponibles en el archivo Excel
wb = load_workbook(xls_path, data_only=True)
sheet_names = wb.sheetnames  # Obtener los nombres de las hojas

# Selecci√≥n de la hoja a cargar (Perros, Gatos, etc.)
sheet_name = st.selectbox("Selecciona la l√≠nea de productos:", sheet_names)

# Cargar los productos desde la hoja seleccionada
df_base, images = load_products_from_sheet(str(xls_path), sheet_name)

if df_base.empty:
    st.stop()  # Detener la ejecuci√≥n si no se encuentran datos v√°lidos

# B√∫squeda en productos
search_term = st.text_input("üîç Buscar (SKU o descripci√≥n)‚Ä¶").strip().lower()
search_norm = quitar_acentos(search_term)

if search_term:
    df = df_base[
        df_base["descripcion_norm"].str.contains(search_norm, na=False)
        | df_base["SKU"].str.contains(search_norm, na=False)
    ]
else:
    df = df_base.copy()

# --- Carrito de compras --- 

if "cart" not in st.session_state:
    st.session_state["cart"] = []

# Funci√≥n para agregar al carrito
def add_to_cart(sku, nombre, precio, cantidad=1):
    st.session_state["cart"].append({"SKU": sku, "Nombre": nombre, "Precio": precio, "Cantidad": cantidad})

# Mostrar carrito y total
def show_cart():
    if st.session_state["cart"]:
        st.write("### Carrito de compras")
        cart_df = pd.DataFrame(st.session_state["cart"])
        st.write(cart_df)
        total = sum(item["Precio"] * item["Cantidad"] for item in st.session_state["cart"])
        st.write(f"**Total: ${total:,.2f}**")
    else:
        st.write("### El carrito est√° vac√≠o.")

# --- Funci√≥n para generar PDF --- 

def generate_pdf(cart_items):
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)
    c.setFont("Helvetica", 12)
    
    c.drawString(30, 750, "Pedido - Cat√°logo Millex")
    c.drawString(30, 735, f"Fecha: {pd.Timestamp.now().strftime('%d/%m/%Y')}")
    
    y_position = 710
    line_height = 15  # Espacio entre las l√≠neas de texto
    for item in cart_items:
        c.drawString(30, y_position, f"SKU: {item['SKU']}")
        c.drawString(100, y_position, f"Nombre: {item['Nombre']}")
        c.drawString(300, y_position, f"Precio: ${item['Precio']}")
        y_position -= line_height
        
        if y_position < 100:
            c.showPage()  # Si estamos cerca del final de la p√°gina, agregamos una nueva
            c.setFont("Helvetica", 12)
            y_position = 750
    
    total = sum(item["Precio"] * item["Cantidad"] for item in cart_items)
    c.drawString(30, y_position - 20, f"**Total del Pedido: ${total:,.2f}**")
    
    c.save()
    buffer.seek(0)
    return buffer

# --- Paginaci√≥n simple ---

ITEMS_PER_PAGE = 20
total_pages = max(1, math.ceil(len(df) / ITEMS_PER_PAGE))
page_key = f"page_{sheet_name}"
if page_key not in st.session_state:
    st.session_state[page_key] = 1

def prev_page():
    if st.session_state[page_key] > 1:
        st.session_state[page_key] -= 1

def next_page():
    if st.session_state[page_key] < total_pages:
        st.session_state[page_key] += 1

# Mostrar paginaci√≥n
col1, col2, col3 = st.columns([1, 3, 1])
with col1:
    st.button("‚óÄ Anterior", on_click=prev_page, disabled=st.session_state[page_key] == 1)
with col2:
    st.markdown(f"**P√°gina {st.session_state[page_key]} de {total_pages}**")
with col3:
    st.button("Siguiente ‚ñ∂", on_click=next_page, disabled=st.session_state[page_key] == total_pages)

start_idx = (st.session_state[page_key] - 1) * ITEMS_PER_PAGE
end_idx = start_idx + ITEMS_PER_PAGE
df_page = df.iloc[start_idx:end_idx]

# --- Mostrar productos con imagen y carrito ---

for _, row in df_page.iterrows():
    st.write(f"**SKU:** {row['SKU']}")
    st.write(f"**Descripci√≥n:** {row['Descripcion']}")
    st.write(f"**Tama√±o del producto:** {row['Tama√±o del producto']}")
    st.write(f"**Precio unitario (USD):** ${row['Precio USD']:,.2f}")
    st.write(f"**Unidades por caja:** {row['Unidades por caja']}")
    
    # Mostrar imagen (si existe en el diccionario de im√°genes)
    if row['SKU'] in images:
        img_bytes = images[row['SKU']]
        st.image(img_bytes, caption=row['Descripcion'], width=150)  # Ajustar tama√±o de imagen a 150px de ancho
    
    # Campo para ingresar cantidad de cajas
    cantidad = st.number_input(f"Cantidad de {row['SKU']}", min_value=1, max_value=100, value=1, step=1, key=row['SKU'])
    
    # Calcular el precio total basado en la cantidad de cajas y el precio por unidad
    total_price = row["Precio USD"] * row["Unidades por caja"] * cantidad
    st.write(f"**Precio Total (USD):** ${total_price:,.2f}")
    
    # Bot√≥n para agregar al carrito
    if st.button(f"Agregar {row['SKU']} al carrito", key=f"add_{row['SKU']}"):
        add_to_cart(row['SKU'], row['Descripcion'], total_price, cantidad)
    
    st.markdown("---")

# Mostrar el carrito
show_cart()

# Bot√≥n para descargar el carrito como PDF
if st.button("Descargar PDF del Pedido"):
    pdf_buffer = generate_pdf(st.session_state["cart"])
    st.download_button(
        label="Descargar Pedido PDF",
        data=pdf_buffer,
        file_name="pedido_cat√°logo_millex.pdf",
        mime="application/pdf"
    )


